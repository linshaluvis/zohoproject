#Zoho Final
from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout
from django.contrib import messages
from django.conf import settings
from datetime import date
from datetime import datetime, timedelta
from Company_Staff.models import *
from django.db import models
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from django.template.loader import get_template
from bs4 import BeautifulSoup
import io,os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse,HttpResponseRedirect
from io import BytesIO
from django.db.models import Max
from django.db.models import Q
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMultiAlternatives


# Create your views here.



# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Calculate the date 20 days before the end date for payment term renew
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        alert_message = current_date >= reminder_date
        
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False

        # Calculate the number of days between the reminder date and end date
        days_left = (dash_details.End_date - current_date).days
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'alert_message':alert_message,
            'days_left':days_left,
            'payment_request':payment_request,
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company staff request for login approval
def company_staff_request(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
       
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request, pk):
    """
    Sets the company approval status to 2 for the specified staff member, effectively canceling staff approval.

    This function is designed to be used for canceling staff approval, and the company approval value is set to 2.
    This can be useful for identifying resigned staff under the company in the future.

    """
    staff = StaffDetails.objects.get(id=pk)
    staff.company_approval = 2
    staff.save()
    return redirect('company_all_staff')


# company profile, profile edit
def company_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        terms=PaymentTerms.objects.all()
        payment_history=dash_details.previous_plans.all()

        # Calculate the date 20 days before the end date
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        renew_button = current_date >= reminder_date

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'renew_button': renew_button,
            'terms':terms,
            'payment_history':payment_history,
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

def company_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile_editpage.html', context)
    else:
        return redirect('/')

def company_profile_basicdetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            messages.success(request,'Updated')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
    
def company_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('company_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
       
def company_profile_companydetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            gstno = request.POST.get('gstno')
            profile_pic = request.FILES.get('image')

            # Update the CompanyDetails object with form data
            dash_details.company_name = request.POST.get('cname')
            dash_details.contact = request.POST.get('phone')
            dash_details.address = request.POST.get('address')
            dash_details.city = request.POST.get('city')
            dash_details.state = request.POST.get('state')
            dash_details.country = request.POST.get('country')
            dash_details.pincode = request.POST.get('pincode')
            dash_details.pan_number = request.POST.get('pannumber')

            if gstno:
                dash_details.gst_no = gstno

            if profile_pic:
                dash_details.profile_pic = profile_pic

            dash_details.save()

            messages.success(request, 'Updated')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/')    

# company modules editpage
def company_module_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_module_editpage.html', context)
    else:
        return redirect('/')

def company_module_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Check for any previous module update request
        if ZohoModules.objects.filter(company=dash_details,status='Pending').exists():
            messages.warning(request,'You have a pending update request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            # Retrieve values
            items = request.POST.get('items', 0)
            price_list = request.POST.get('price_list', 0)
            stock_adjustment = request.POST.get('stock_adjustment', 0)
            godown = request.POST.get('godown', 0)

            cash_in_hand = request.POST.get('cash_in_hand', 0)
            offline_banking = request.POST.get('offline_banking', 0)
            upi = request.POST.get('upi', 0)
            bank_holders = request.POST.get('bank_holders', 0)
            cheque = request.POST.get('cheque', 0)
            loan_account = request.POST.get('loan_account', 0)

            customers = request.POST.get('customers', 0)
            invoice = request.POST.get('invoice', 0)
            estimate = request.POST.get('estimate', 0)
            sales_order = request.POST.get('sales_order', 0)
            recurring_invoice = request.POST.get('recurring_invoice', 0)
            retainer_invoice = request.POST.get('retainer_invoice', 0)
            credit_note = request.POST.get('credit_note', 0)
            payment_received = request.POST.get('payment_received', 0)
            delivery_challan = request.POST.get('delivery_challan', 0)

            vendors = request.POST.get('vendors', 0)
            bills = request.POST.get('bills', 0)
            recurring_bills = request.POST.get('recurring_bills', 0)
            vendor_credit = request.POST.get('vendor_credit', 0)
            purchase_order = request.POST.get('purchase_order', 0)
            expenses = request.POST.get('expenses', 0)
            recurring_expenses = request.POST.get('recurring_expenses', 0)
            payment_made = request.POST.get('payment_made', 0)

            projects = request.POST.get('projects', 0)

            chart_of_accounts = request.POST.get('chart_of_accounts', 0)
            manual_journal = request.POST.get('manual_journal', 0)

            eway_bill = request.POST.get('ewaybill', 0)

            employees = request.POST.get('employees', 0)
            employees_loan = request.POST.get('employees_loan', 0)
            holiday = request.POST.get('holiday', 0)
            attendance = request.POST.get('attendance', 0)
            salary_details = request.POST.get('salary_details', 0)

            reports = request.POST.get('reports', 0)

            update_action=1
            status='Pending'

            # Create a new ZohoModules instance and save it to the database
            data = ZohoModules(
                company=dash_details,
                items=items, price_list=price_list, stock_adjustment=stock_adjustment, godown=godown,
                cash_in_hand=cash_in_hand, offline_banking=offline_banking, upi=upi, bank_holders=bank_holders,
                cheque=cheque, loan_account=loan_account,
                customers=customers, invoice=invoice, estimate=estimate, sales_order=sales_order,
                recurring_invoice=recurring_invoice, retainer_invoice=retainer_invoice, credit_note=credit_note,
                payment_received=payment_received, delivery_challan=delivery_challan,
                vendors=vendors, bills=bills, recurring_bills=recurring_bills, vendor_credit=vendor_credit,
                purchase_order=purchase_order, expenses=expenses, recurring_expenses=recurring_expenses,
                payment_made=payment_made,
                projects=projects,
                chart_of_accounts=chart_of_accounts, manual_journal=manual_journal,
                eway_bill=eway_bill,
                employees=employees, employees_loan=employees_loan, holiday=holiday,
                attendance=attendance, salary_details=salary_details,
                reports=reports,update_action=update_action,status=status    
            )
            data.save()
            messages.success(request,"Request sent successfully. Please wait for approval.")
            return redirect('company_profile')
        else:
            return redirect('company_module_editpage')  
    else:
        return redirect('/')


def company_renew_terms(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        # Check for any previous  extension request
        if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists():
            messages.warning(request,'You have a pending request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            select=request.POST['select']
            terms=PaymentTerms.objects.get(id=select)
            update_action=1
            status='Pending'
            newterms=PaymentTermsUpdates(
               company=dash_details,
               payment_term=terms,
               update_action=update_action,
               status=status 
            )
            newterms.save()
            messages.success(request,'Request sent successfully, Please wait for approval...')
            return redirect('company_profile')
        else:
            return redirect('company_profile')
    else:
        return redirect('/')

# company notifications and messages
def company_notifications(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        notifications = dash_details.notifications.filter(is_read=0).order_by('-date_created','-time')
        end_date = dash_details.End_date
        company_days_remaining = (end_date - date.today()).days
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False
        
        print(company_days_remaining)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'notifications':notifications,
            'days_remaining':company_days_remaining,
            'payment_request':payment_request,
        }

        return render(request,'company/company_notifications.html',context)
        
    else:
        return redirect('/')
        
        
def company_message_read(request,pk):
    '''
    message read functions set the is_read to 1, 
    by default it is 0 means not seen by user.

    '''
    notification=Notifications.objects.get(id=pk)
    notification.is_read=1
    notification.save()
    return redirect('company_notifications')
    
    
def company_payment_history(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        payment_history=dash_details.previous_plans.all()

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'payment_history':payment_history,
            
        }
        return render(request,'company/company_payment_history.html', context)
    else:
        return redirect('/')
        
def company_trial_feedback(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        trial_instance = TrialPeriod.objects.get(company=dash_details)
        if request.method == 'POST':
            interested = request.POST.get('interested')
            feedback=request.POST.get('feedback') 
            
            trial_instance.interested_in_buying=1 if interested =='yes' else 2
            trial_instance.feedback=feedback
            trial_instance.save()

            if interested =='yes':
                return redirect('company_profile')
            else:
                return redirect('company_dashboard')
        else:
            return redirect('company_dashboard')
    else:
        return redirect('/')
# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')


def staff_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'staff/staff_profile_editpage.html', context)
    else:
        return redirect('/')

def staff_profile_details_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            dash_details.contact = request.POST.get('phone')
            old=dash_details.image
            new=request.FILES.get('profile_pic')
            print(new,old)
            if old!=None and new==None:
                dash_details.image=old
            else:
                print(new)
                dash_details.image=new
            dash_details.save()
            messages.success(request,'Updated')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

def staff_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('staff_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')
    
    
# staff invoice section
def invoice_list_out(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
       

        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)


        return render(request,'staff/invoicelist.html',{'allmodules':allmodules,'data':log_details,'details':dash_details,'invoices':invoices})
    else:
       return redirect('/')

   
def view(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
       
        

        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)

   
        inv = invoice.objects.get(id = pk)
        cmt = invoicecomments.objects.filter(invoice = inv)
        hist =invoiceHistory.objects.filter( invoice = inv).last()
        histo =invoiceHistory.objects.filter( invoice = inv)

        invItems = invoiceitems.objects.filter( invoice = inv)
        created = invoiceHistory.objects.filter( invoice = inv,  action = 'Created')
        price_lists=PriceList.objects.filter(company=company,status='Active')

        
        return render(request,'staff/invoice.html',{'allmodules':allmodules,'com':company,'cmp':company, 'data':log_details, 'details': dash_details,'invoice':inv,'invoices':invoices,'invItems':invItems, 'comments':cmt,'history':hist,'historys':histo,  'created':created,'price_lists':price_lists})
    else:
       return redirect('/')

def convertInvoice(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)



        inv = invoice.objects.get(id = id)
        inv.status = 'Saved'
        inv.save()
        return redirect(view,id)
    
def add_attach(request,id):
    if request.method == 'POST' and request.FILES.get('file'):
        inv = invoice.objects.get(id=id)
        inv.document = request.FILES['file']
        print("success")

        inv.save()

        
        return redirect(view, id)
    
def invoicePdf(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)


        
        inv = invoice.objects.get(id = id)
        itms = invoiceitems.objects.filter(invoice = inv)
    
        context = {'invoice':inv, 'invItems':itms,'cmp':company}
        
        template_path = 'staff/invoice_Pdf.html'
        fname = 'Invoice_'+inv.invoice_number
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] =f'attachment; filename = {fname}.pdf'
        # find the template and render it.
        template = get_template(template_path)
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(
        html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response
    else:
        return redirect('view')
    
def InvoiceHistory(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)

        inv = invoice.objects.get(id = id)
        his = invoiceHistory.objects.filter(invoice = inv)
       
        
        return render(request,'staff/invoice_History.html',{'allmodules':allmodules,'com':company,'data':log_details,'history':his, 'invoice':inv})
    else:
       return redirect('/')
def deleteInvoice(request, id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)

        inv = invoice.objects.get( id = id)


  

        for i in invoiceitems.objects.filter(invoice = inv):
            item = Items.objects.get(id = i.Items.id)
            print(item)
            item.current_stock += i.quantity
            item.save()
        
        invoiceitems.objects.filter(invoice = inv).delete()

        # Storing ref number to deleted table
        # if entry exists and lesser than the current, update and save => Only one entry per company
        if invoiceReference.objects.filter(company = company).exists():
            deleted = invoiceReference.objects.get(company = company)
            if int(inv.reference_number) > int(deleted.reference_number):
                deleted.reference_number = inv.reference_number
                deleted.save()
        else:
            invoiceReference.objects.create(company = company, reference_number = inv.reference_number)
        
        inv.delete()
        return redirect(invoice_list_out)
def editInvoice(request,id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company
            dash_details = StaffDetails.objects.get(login_details=log_details)

        allmodules= ZohoModules.objects.get(company = cmp)
        cust = Customer.objects.filter(company = cmp, customer_status = 'Active')
        trm = Company_Payment_Term.objects.filter(company = cmp)
        repeat = CompanyRepeatEvery.objects.filter(company = cmp)
        bnk = Banking.objects.filter(company = cmp)
        priceList = PriceList.objects.filter(company = cmp, status = 'Active')
        itms = Items.objects.filter(company = cmp, activation_tag = 'active')
        units = Unit.objects.filter(company=cmp)
        accounts=Chart_of_Accounts.objects.filter(company=cmp)

        invoices = invoice.objects.get(id = id)
        invItems = invoiceitems.objects.filter(invoice = invoices)

        context = {
            'cmp':cmp,'allmodules':allmodules, 'details':dash_details, 'customers': cust,'pTerms':trm, 'repeat':repeat, 'banks':bnk, 'priceListItems':priceList, 'items':itms,
            'units': units,'accounts':accounts, 'invoice':invoices, 'invItems': invItems,
        }
    
        
        return render(request,'staff/edit_invoice.html',context)
    else:
       return redirect('/')
def updateInvoice(request, id):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        inv = invoice.objects.get(id = id)
        if request.method == 'POST':
            invNum = request.POST['rec_invoice_no']
            if inv.invoice_number != invNum and invoice.objects.filter(company = com, invoice_number__iexact = invNum).exists():

                res = f'<script>alert(" Invoice Number `{invNum}` already exists, try another!");window.history.back();</script>'
                return HttpResponse(res)

            inv.customer = Customer.objects.get(id = request.POST['customerId'])
            inv.customer_email = request.POST['customer_email']
            inv.customer_billingaddress = request.POST['bill_address']
            inv.customer_GSTtype = request.POST['customer_gst_type']
            y=request.POST['customer_gst_type']
            print(y)
            inv.customer_GSTnumber = request.POST['customer_gstin']
            inv.customer_place_of_supply = request.POST['place_of_supply']
            x=request.POST['place_of_supply']
            print(x)
           
            inv.reference_number = request.POST['reference_number']
            inv.invoice_number = invNum
            inv.payment_terms = Company_Payment_Term.objects.get(id = request.POST['payment_term'])
            inv.date = request.POST['start_date']
            inv.expiration_date = datetime.strptime(request.POST['end_date'], '%d-%m-%Y').date()
            # inv.order_no = request.POST['order_number']
            inv.price_list_applied = True if 'priceList' in request.POST else False
            inv.price_list = None if request.POST['price_list_id'] == "" else PriceList.objects.get(id = request.POST['price_list_id'])
            # inv.repeat_every = CompanyRepeatEvery.objects.get(id = request.POST['repeat_every'])
            inv.payment_method = None if request.POST['payment_method'] == "" else request.POST['payment_method']
            inv.cheque_number = None if request.POST['cheque_id'] == "" else request.POST['cheque_id']
            inv.UPI_number = None if request.POST['upi_id'] == "" else request.POST['upi_id']
            inv.bank_account_number = None if request.POST['bnk_id'] == "" else request.POST['bnk_id']
            inv.sub_total = 0.0 if request.POST['subtotal'] == "" else float(request.POST['subtotal'])
            inv.IGST = 0.0 if request.POST['igst'] == "" else float(request.POST['igst'])
            inv.CGST = 0.0 if request.POST['cgst'] == "" else float(request.POST['cgst'])
            inv.SGST = 0.0 if request.POST['sgst'] == "" else float(request.POST['sgst'])
            inv.tax_amount = 0.0 if request.POST['taxamount'] == "" else float(request.POST['taxamount'])
            inv.adjustment = 0.0 if request.POST['adj'] == "" else float(request.POST['adj'])
            inv.shipping_charge = 0.0 if request.POST['ship'] == "" else float(request.POST['ship'])
            inv.grand_total = 0.0 if request.POST['grandtotal'] == "" else float(request.POST['grandtotal'])
            inv.advanced_paid = 0.0 if request.POST['advance'] == "" else float(request.POST['advance'])
            inv.balance = request.POST['grandtotal'] if request.POST['balance'] == "" else float(request.POST['balance'])
            inv.description = request.POST['note']
            inv.terms_and_condition = request.POST['terms']

            if len(request.FILES) != 0:
                inv.document=request.FILES.get('file')
            inv.save()


            # Save rec_invoice items.

            itemId = request.POST.getlist("item_id[]")
            itemName = request.POST.getlist("item_name[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("priceListPrice[]") if 'priceList' in request.POST else request.POST.getlist("price[]")
            tax = request.POST.getlist("taxGST[]") if request.POST['place_of_supply'] == com.state else request.POST.getlist("taxIGST[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")
            inv_item_ids = request.POST.getlist("id[]")
            invItem_ids = [int(id) for id in inv_item_ids]

            inv_items = invoiceitems.objects.filter(invoice = inv)
            object_ids = [obj.id for obj in inv_items]

            ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in invItem_ids]
            for itmId in ids_to_delete:
                invItem = invoiceitems.objects.get(id = itmId)
                item = Items.objects.get(id = invItem.item.id)
                item.current_stock += invItem.quantity
                item.save()

            invoiceitems.objects.filter(id__in=ids_to_delete).delete()
            
            count = invoiceitems.objects.filter(invoice = inv).count()

            if len(itemId)==len(itemName)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total)==len(invItem_ids) and invItem_ids and itemId and itemName and hsn and qty and price and tax and discount and total:
                mapped = zip(itemId,itemName,hsn,qty,price,tax,discount,total,invItem_ids)
                mapped = list(mapped)
                for ele in mapped:
                    if int(len(itemId))>int(count):
                        if ele[8] == 0:
                            itm = Items.objects.get(id = int(ele[0]))
                            invoiceitems.objects.create(company = com, logindetails = com.login_details, invoice = inv, Items = itm, hsn = ele[2], quantity = int(ele[3]), price = float(ele[4]), tax_rate = ele[5], discount = float(ele[6]), total = float(ele[7]))
                            itm.current_stock -= int(ele[3])
                            itm.save()
                        else:
                            itm = Items.objects.get(id = int(ele[0]))
                            inItm = invoiceitems.objects.get(id = int(ele[8]))
                            crQty = int(inItm.quantity)
                            
                            invoiceitems.objects.filter( id = int(ele[8])).update(invoice = inv, Items = itm, hsn = ele[2], quantity = int(ele[3]), price = float(ele[4]), tax_rate = ele[5], discount = float(ele[6]), total = float(ele[7]))

                            if crQty < int(ele[3]):
                                itm.current_stock -=  abs(crQty - int(ele[3]))
                            elif crQty > int(ele[3]):
                                itm.current_stock += abs(crQty - int(ele[3]))
                            itm.save()
                    else:
                        itm = Items.objects.get(id = int(ele[0]))
                        inItm = invoiceitems.objects.get(id = int(ele[8]))
                        crQty = int(inItm.quantity)

                        invoiceitems.objects.filter( id = int(ele[8])).update(invoice = inv, Items = itm, hsn = ele[2], quantity = int(ele[3]), price = float(ele[4]), tax_rate = ele[5], discount = float(ele[6]), total = float(ele[7]))

                        if crQty < int(ele[3]):
                            itm.current_stock -=  abs(crQty - int(ele[3]))
                        elif crQty > int(ele[3]):
                            itm.current_stock += abs(crQty - int(ele[3]))
                        itm.save()
            
            # Save transaction
                    
            invoiceHistory.objects.create(
                company = com,
                login_details = com.login_details,
                invoice = inv,
                action = 'Edited'
            )

            return redirect(view, id)
        else:
            return redirect(editInvoice, id)
    else:
       return redirect('/')
    
def filter_invoice_name(request, pk):
    if 'login_id' not in request.session:
                return redirect('/')
    log_id = request.session['login_id']

    log_details= LoginDetails.objects.get(id=log_id)
    dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    log_id = request.session['login_id']
    log_details = LoginDetails.objects.get(id=log_id)
    cmp =dash_details.company


    if log_details.user_type == 'Staff':
        staff = StaffDetails.objects.get(login_details=log_details)
        company = staff.company
    elif log_details.user_type == 'Company':
        company = CompanyDetails.objects.get(login_details=log_details)
    
    try:
        invoic = invoice.objects.get(id=pk)
        item = invoiceitems.objects.filter(invoice=pk)
        customers = Customer.objects.filter(company_id=company, customer_status='Active')
        print(customers)
        cmt = invoicecomments.objects.filter(invoice = invoic)
        hist =invoiceHistory.objects.filter( invoice = invoic).last()
        histo =invoiceHistory.objects.filter( invoice = invoic)

        invItems = invoiceitems.objects.filter( invoice = invoic)
        created = invoiceHistory.objects.filter( invoice = invoic,  action = 'Created')


        for r in customers:
            vn = r.first_name.split()[1:]
            r.cust_name = " ".join(vn)

        sorted_customers = sorted(customers, key=lambda r: r.cust_name)
        print(sorted_customers)
        for customer in sorted_customers:
            print(customer.first_name)  # Assuming you have a field named 'cust_name'


        context = {
            'allmodules':allmodules,
            'com':company,
            'data':log_details, 
            'details': dash_details,
            'invoices': sorted_customers,
            'invoice': invoic,
            'invItems': invItems,
            'company': company,
            'comments':cmt,
            'history':hist,
            'historys':histo, 
            'created':created,
            'cmp':cmp
        }
        return render(request, 'staff/invoice.html', context)
    
    except invoice.DoesNotExist:
        return redirect('/')
def filter_invoice_number(request, pk):
    if 'login_id' not in request.session:
        return redirect('/')
    log_id = request.session['login_id']

    log_details= LoginDetails.objects.get(id=log_id)
    dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')    
    log_id = request.session['login_id']
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == 'Staff':
        staff = StaffDetails.objects.get(login_details=log_details)
        company = staff.company
    elif log_details.user_type == 'Company':
        company = CompanyDetails.objects.get(login_details=log_details)
    
    try:
        invoic = invoice.objects.get(id=pk)
        item = invoiceitems.objects.filter(invoice=pk)
        invoices = invoice.objects.filter(company=company)
        
        for r in invoices:
            vn = r.invoice_number.split()[1:]  # accessing attributes using dot notation
            r.cust_no = " ".join(vn)

        sorted_invoices = sorted(invoices, key=lambda r: r.cust_no)
        for customer in sorted_invoices:
            print(customer.invoice_number) 
        

        context = {
             'allmodules':allmodules,
            'com':company,
            'data':log_details, 
            'details': dash_details,
            'invoices': sorted_invoices,
            'invoice': invoic,
            'item': item,
            'company': company,
        }
        return render(request, 'staff/invoice.html', context)
    
    except invoice.DoesNotExist:
        return redirect('/') 
def addInvoiceComment(request, id):
    if 'login_id' not in request.session:
        return redirect('/')
    
    log_id = request.session['login_id']
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
            com = CompanyDetails.objects.get(login_details=log_details)
    else:
            cmp = StaffDetails.objects.get(login_details=log_details)
            com = cmp.company

            

    inv = invoice.objects.get(id = id)
    if request.method == "POST":
            cmt = request.POST['comment'].strip()

            invoicecomments.objects.create(company = com, invoice = inv, comments = cmt)
            return redirect(view, id)
    return redirect(view, id)
def deleteInvoiceComment(request,id):
    if 'login_id' not in request.session:
        return redirect('/')
    print(id)
    cmt = invoicecomments.objects.get(id = id)
    invId = cmt.invoice.id
    cmt.delete()

    return redirect(view,invId)

def shareInvoiceToEmail(request,id):
    if 'login_id' not in request.session:
        return redirect('/')
    
    log_id = request.session['login_id']
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == 'Staff':
        staff = StaffDetails.objects.get(login_details=log_details)
        company = staff.company
    elif log_details.user_type == 'Company':
        company = CompanyDetails.objects.get(login_details=log_details)
        
    inv = invoice.objects.get(id = id)
    itms = invoiceitems.objects.filter(invoice = inv)
    try:
        if request.method == 'POST':
            emails_string = request.POST['email_ids']
            print(emails_string)

            emails_list = [email.strip() for email in emails_string.split(',')]
            email_message = request.POST['email_message']
            
        
            context = {'invoice':inv, 'invItems':itms,'cmp':company}
            template_path = 'staff/invoice_pdf.html'
            template = get_template(template_path)

            

            html  = template.render(context)
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
            pdf = result.getvalue()
            filename = f'Invoice_{inv.invoice_number}'
            subject = f"Invoice_{inv.invoice_number}"

            email = EmailMultiAlternatives(subject, f"Hi,\nPlease find the attached Invoice for - INVOICE-{inv.invoice_number}. \n{email_message}\n\n--\nRegards,\n{company.company_name}\n{company.address}\n{company.state} - {company.country}\n{company.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
            email.attach(filename, pdf, "application/pdf")
            email.send(fail_silently=False)

            messages.success(request, 'Invoice details has been shared via email successfully..!')
            return redirect(view,id)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(view, id)
def filter_invoice_draft(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        invo=invoice.objects.filter(status='draft',company = company)
        invoic=invoice.objects.get(id=pk)
        item=invoiceitems.objects.filter(invoice=pk)

        context={'invoices':invo,'invoice':invoic,'item':item}
        return render(request,'staff/invoice.html',context)
    
    
def filter_invoice_sent(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        invo=invoice.objects.filter(status='saved',company = company)
        invoic=invoice.objects.get(id=pk)
        item=invoiceitems.objects.filter(invoice=pk)

        context={'invoices':invo,'invoice':invoic,'item':item}
        return render(request,'staff/invoice.html',context)    
def invoice_create(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        customers=Customer.objects.all()
        item=Items.objects.all()
        payments=Company_Payment_Term.objects.all()
        i = invoice.objects.all()
        acc = Chart_of_Accounts.objects.filter(Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')).order_by('account_name')


        if invoice.objects.all().exists():
            invoice_count = invoice.objects.last().id
            count = invoice_count
        else:
            count = 1



       
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'customers':customers,
            'item':item,
            'payments':payments,
            'count': count,
            'i':i,
            'accounts':acc,
            'company':company,
            


            
        }
        return render(request,'staff/invoice.html',context)  
    
def invoice_createpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company
            dash_details = StaffDetails.objects.get(login_details=log_details)

        allmodules= ZohoModules.objects.get(company = cmp)
        cust = Customer.objects.filter(company = cmp, customer_status = 'Active')
        trm = Company_Payment_Term.objects.filter(company = cmp)
        repeat = CompanyRepeatEvery.objects.filter(company = cmp)
        bnk = Banking.objects.filter(company = cmp)
        priceList = PriceList.objects.filter(company = cmp, status = 'Active')
        itms = Items.objects.filter(company = cmp, activation_tag = 'active')
        
        units = Unit.objects.filter(company=cmp)
        accounts=Chart_of_Accounts.objects.filter(company=cmp)

        # Fetching last rec_invoice and assigning upcoming ref no as current + 1
        # Also check for if any bill is deleted and ref no is continuos w r t the deleted rec_invoice
        latest_inv = invoice.objects.filter(company = cmp).order_by('-id').first()

        new_number = int(latest_inv.reference_number) + 1 if latest_inv else 1

        if invoiceReference.objects.filter(company = cmp).exists():
            deleted = invoiceReference.objects.get(company = cmp)
            
            if deleted:
                while int(deleted.reference_number) >= new_number:
                    new_number+=1

        # Finding next rec_invoice number w r t last rec_invoice number if exists.
        nxtInv = ""
        lastInv = invoice.objects.filter(company = cmp).last()
        if lastInv:
            inv_no = str(lastInv.invoice_number)
            numbers = []
            stri = []
            for word in inv_no:
                if word.isdigit():
                    numbers.append(word)
                else:
                    stri.append(word)
            
            num=''
            for i in numbers:
                num +=i
            
            st = ''
            for j in stri:
                st = st+j

            inv_num = int(num)+1

            padding_length = len(num) - 1

                    
            nxtInv = f"{st}{num[0]}{inv_num:0{padding_length}d}"
        else:
            nxtInv = 'in-01'
        context = {
            'cmp':cmp,'allmodules':allmodules, 'details':dash_details, 'customers': cust,'pTerms':trm, 'repeat':repeat, 'banks':bnk, 'priceListItems':priceList, 'items':itms,
            'invNo':nxtInv, 'ref_no':new_number,'units': units,'accounts':accounts,
        }
    
        return render(request,'staff/createinvoice.html',context)
    
def viewInvoice(request):
  
    customer_id = request.GET.get('cust')
    cust = Customer.objects.get(id=customer_id)

    try:
        id = request.GET.get('id')

        try:
            item = Items.objects.get(item_name=id)
            name = item.item_name
            rate = item.selling_price
            hsn = item.hsn_code
            avl=item.current_stock
            # Assuming `company_name` is a field in the `company_details` model
            place = cust.place_of_supply
            return JsonResponse({"status": "not", 'place': place, 'rate': rate, 'avl':avl ,'hsn': hsn})
        except Items.DoesNotExist:
            return JsonResponse({"status": "error", 'message': "Item not found"})
    except Exception as e:
        return JsonResponse({"status": "error", 'message': str(e)})
def customerdata(request):
    customer_id = request.GET.get('id')
    print(customer_id)
    cust = Customer.objects.get(id=customer_id)
    data7 = {'email': cust.customer_email,'gst':cust.GST_treatement,'gstin':cust.GST_number}
    
    return JsonResponse(data7)
def getInvItemDetails(request):
  
       
        
        itemName = request.GET['item']
        item = Items.objects.get( item_name = itemName)

        context = {
            'status':True,
            'id':item.id,
            'hsn':item.hsn_code,
            'sales_rate':item.selling_price,
            'avl':item.current_stock,
            'tax': True if item.tax_reference == 'taxable' else False,
            'gst':item.intrastate_tax,
            'igst':item.interstate_tax,

        }
        return JsonResponse(context)
def getBankAccount(request):
  
        
       bankId = request.GET['id']
       bnk = Banking.objects.get(id = bankId)

       if bnk:
            return JsonResponse({'status':True, 'account':bnk.bnk_acno})
       else:
            return JsonResponse({'status':False, 'message':'Something went wrong..!'})
def createInvoice(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        if request.method == 'POST':
            invNum = request.POST['rec_invoice_no']
            if invoice.objects.filter(company = com, invoice_number__iexact = invNum).exists():
                res = f'<script>alert("Invoice Number `{invNum}` already exists, try another!");window.history.back();</script>'
                return HttpResponse(res)

            inv = invoice(
                company = com,
                login_details = com.login_details,
                customer = Customer.objects.get(id = request.POST['customerId']),
                customer_email = request.POST['customer_email'],
                customer_billingaddress = request.POST['bill_address'],
                customer_GSTtype = request.POST['customer_gst_type'],
                customer_GSTnumber = request.POST['customer_gstin'],
                customer_place_of_supply = request.POST['place_of_supply'],
               
                reference_number = request.POST['reference_number'],
                invoice_number = invNum,
                payment_terms = Company_Payment_Term.objects.get(id = request.POST['payment_term']),
                date = request.POST['start_date'],
                expiration_date = datetime.strptime(request.POST['end_date'], '%d-%m-%Y').date(),
                # Order_number = request.POST['order_number'],
                price_list_applied = True if 'priceList' in request.POST else False,
                price_list = None if request.POST['price_list_id'] == "" else PriceList.objects.get(id = request.POST['price_list_id']),
                payment_method = None if request.POST['payment_method'] == "" else request.POST['payment_method'],
                cheque_number = None if request.POST['cheque_id'] == "" else request.POST['cheque_id'],
                UPI_number = None if request.POST['upi_id'] == "" else request.POST['upi_id'],
                bank_account_number = None if request.POST['bnk_id'] == "" else request.POST['bnk_id'],
                sub_total = 0.0 if request.POST['subtotal'] == "" else float(request.POST['subtotal']),
                IGST = 0.0 if request.POST['igst'] == "" else float(request.POST['igst']),
                CGST = 0.0 if request.POST['cgst'] == "" else float(request.POST['cgst']),
                SGST = 0.0 if request.POST['sgst'] == "" else float(request.POST['sgst']),
                tax_amount = 0.0 if request.POST['taxamount'] == "" else float(request.POST['taxamount']),
                adjustment = 0.0 if request.POST['adj'] == "" else float(request.POST['adj']),
                shipping_charge = 0.0 if request.POST['ship'] == "" else float(request.POST['ship']),
                grand_total = 0.0 if request.POST['grandtotal'] == "" else float(request.POST['grandtotal']),
                advanced_paid = 0.0 if request.POST['advance'] == "" else float(request.POST['advance']),
                balance = request.POST['grandtotal'] if request.POST['balance'] == "" else float(request.POST['balance']),
                description = request.POST['note'],
                terms_and_condition = request.POST['terms']
            )

            inv.save()

            if len(request.FILES) != 0:
                inv.document=request.FILES.get('file')
            inv.save()

            if 'Draft' in request.POST:
                inv.status = "Draft"
            elif "Saved" in request.POST:
                inv.status = "Saved" 
            inv.save()

            # Save rec_invoice items.

            itemId = request.POST.getlist("item_id[]")
            itemName = request.POST.getlist("item_name[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("priceListPrice[]") if 'priceList' in request.POST else request.POST.getlist("price[]")
            tax = request.POST.getlist("taxGST[]") if request.POST['place_of_supply'] == com.state else request.POST.getlist("taxIGST[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")

            if len(itemId)==len(itemName)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total) and itemId and itemName and hsn and qty and price and tax and discount and total:
                mapped = zip(itemId,itemName,hsn,qty,price,tax,discount,total)
                mapped = list(mapped)
                for ele in mapped:
                    itm = Items.objects.get(id = int(ele[0]))
                    invoiceitems.objects.create(company = com, logindetails = com.login_details, invoice = inv, Items = itm, hsn = ele[2], quantity = int(ele[3]), price = float(ele[4]), tax_rate = ele[5], discount = float(ele[6]), total = float(ele[7]))
                    itm.current_stock -= int(ele[3])
                    itm.save()

            # Save transaction
                    
            invoiceHistory.objects.create(
                company = com,
                login_details = com.login_details,
                invoice = inv,
                action = 'Created'
            )

            return redirect(invoice_list_out)
        else:
            return redirect(createInvoice)
    else:
       return redirect('/')
   
def invoice_import(request):
    if request.method == 'POST' and 'file' in request.FILES:
        if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
            log_details = LoginDetails.objects.get(id=log_id)

            if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                    
            elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)

            excel_file = request.FILES['file']
            workbook = load_workbook(excel_file)
            
            sheet1 = workbook['Sheet1']
            sheet2 = workbook['Sheet2']
            
            invoices = []  # List to store created invoices

            for row in sheet1.iter_rows(min_row=2, values_only=True):
                try:
                    customer = Customer.objects.get(first_name=row[1],customer_email=row[2],company=company)
                    payment_terms1 = Company_Payment_Term.objects.get(term_name=row[9], company=company)
                    print(payment_terms1)
                except ObjectDoesNotExist:
                    print(f"Customer with name or email '{row[1]}' or Payment term with term name '{row[9]}' does not exist in the database.")
                    continue
                
                # Create and save the invoice object
                latest_inv = invoice.objects.filter(company_id = company).order_by('-id').first()

                new_number = int(latest_inv.reference_number) + 1 if latest_inv else 1

                if invoiceReference.objects.filter(company_id = company).exists():
                    deleted = invoiceReference.objects.get(company_id = company)
                    
                    if deleted:
                        while int(deleted.reference_number) >= new_number:
                            new_number+=1
                created_invoice = invoice(
                    company=company,
                    login_details=log_details,
                    customer=customer,
                    payment_terms=payment_terms1,
                    customer_email=row[2],
                    customer_billingaddress=row[3],
                    customer_GSTtype=row[4],
                    customer_GSTnumber=row[5],
                    customer_place_of_supply=row[6],
                    invoice_number=row[0],
                    date=row[8],
                    expiration_date=row[10],
                    payment_method=row[12],
                    cheque_number=row[13],
                    UPI_number=row[14],
                    bank_account_number=row[15],
                    sub_total=row[19],
                    CGST=row[20],
                    SGST=row[21],
                    IGST=row[29],

                    
                    tax_amount=row[22],
                    shipping_charge=row[23],
                    grand_total=row[25],
                    advanced_paid=row[26],
                    balance=row[27],
                    description=row[16],
                    status=row[28],
                    reference_number=new_number

                )
               
                created_invoice.save()
                invoices.append(created_invoice)
                
                # Save invoice history
                invoiceHistory.objects.create(
                    company=company,
                    login_details=log_details,
                    invoice=created_invoice,
                    date=datetime.now(),
                    action='Created'
                )
        
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                try:
                    item = Items.objects.get(item_name=row[1],company=company)
                except ObjectDoesNotExist:
                    print(f"Item with name '{row[1]}' does not exist in the database.")
                    continue
                
                matching_invoices = [inv for inv in invoices if inv.invoice_number == row[0]]
                if not matching_invoices:
                    print(f"No invoice found for row with invoice number '{row[0]}'")
                    continue

                # Assuming there's only one matching invoice
                invoice1 = matching_invoices[0]

                # Create and save the invoice item object
                invoice_item = invoiceitems(
                    invoice=invoice1,
                    company=company,
                    Items=item,
                    logindetails=log_details,
                    hsn=row[2],
                    quantity=row[3],
                    price=row[4],
                    tax_rate=row[5],
                    discount=row[6],
                    total=row[7],
                )
                invoice_item.save()
                
                # Update current stock for the item
                item.current_stock -= int(row[3])
                item.save()

            return redirect('invoice_list_out')

    return HttpResponse("No file uploaded or invalid request method")

def checkInvoiceNumber(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        
        invNo = request.GET['invNum']
        

        nxtInv = ""
        lastInv = invoice.objects.filter(company = company).last()
        if lastInv:
            inv_no = str(lastInv.invoice_number)

            numbers = []
            stri = []
            for word in inv_no:
                if word.isdigit():
                    numbers.append(word)
                else:
                    stri.append(word)
            
            num=''
            for i in numbers:
                num +=i
            
            st = ''
            for j in stri:
                st = st+j

            inv_num = int(num)+1

            if num[0] == '0':
                if inv_num <10:
                    nxtInv = st+'0'+ str(inv_num)
                else:
                    nxtInv = st+ str(inv_num)
            else:
                nxtInv = st+ str(inv_num)
        if invoice.objects.filter(company = company, invoice_number__iexact = invNo).exists():
            return JsonResponse({'status':False, 'message':'Invoice No already Exists.!'})
        elif nxtInv != "" and invNo != nxtInv:
            return JsonResponse({'status':False, 'message':'Invoice No is not continuous.!'})
        else:
            return JsonResponse({'status':True, 'message':'Number is okay.!'})
   
    

def getInvoiceCustomerData(request):
   
        
        custId = request.POST['id']
        cust = Customer.objects.get(id = custId)

        if cust:
            
                list = False
                listId = None
                listName = None
                context = {
                'status':True, 'id':cust.id, 'email':cust.customer_email, 'gstType':cust.GST_treatement,'shipState':cust.place_of_supply,'gstin':False if cust.GST_number == "" or cust.GST_number == None else True, 'gstNo':cust.GST_number, 'priceList':list, 'ListId':listId, 'ListName':listName,
                'street':cust.billing_address, 'city':cust.billing_city, 'state':cust.billing_state, 'country':cust.billing_country, 'pincode':cust.billing_pincode
                }
                return JsonResponse(context)
def invoiceoverview(request):
    if request.method == 'POST' and 'file' in request.FILES:
        if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        customers=Customer.objects.all()
        item=Items.objects.all()
        payments=Company_Payment_Term.objects.all()
        i = invoice.objects.all()

            



        
        context={
                'details':dash_details,
                'allmodules': allmodules,
                'customers':customers,
                'item':item,
                'payments':payments,
                'i':i


                
            }
        return render(request,'staff/invoice.html',context) 

       
def itemdata(request):
    cur_user = request.user.id
    user = User.objects.get(id=cur_user)
    print(user)

    company = CompanyDetails.objects.get(user = user)
    print(company.state)
    id = request.GET.get('id')
    cust = request.GET.get('cust')
   
        
    item = Items.objects.get(item_name=id)
    cus=Customer.objects.get(first_name=cust)
    rate = item.selling_price
    place=company.state
    gst = item.intrastate_tax
    igst = item.interstate_tax
    desc=item.sales_description
    print(place)
    mail=cus.customer_email
    
    placeof_supply = Customer.objects.get(id=cust).place_of_supply
    print(placeof_supply)
    return JsonResponse({"status":" not",'mail':mail,'desc':desc,'place':place,'rate':rate,'pos':placeof_supply,'gst':gst,'igst':igst})
    return redirect('/')
def checkCustomerName(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        fName = request.POST['first_name']
        lName = request.POST['last_name']

        if Customer.objects.filter(company = company, first_name__iexact = fName, last_name__iexact = lName).exists():
            msg = f'{fName} {lName} already exists, Try another.!'
            return JsonResponse({'is_exist':True, 'message':msg})
        else:
            return JsonResponse({'is_exist':False})
    else:
        return redirect('/')
def checkCustomerGSTIN(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        data = LoginDetails.objects.get(id = log_id)
        log_details= LoginDetails.objects.get(id=log_id)
        
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                company = staff.company
                allmodules=ZohoModules.objects.get(company=staff.company)
                dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
                    
        elif log_details.user_type == 'Company':
                company = CompanyDetails.objects.get(login_details=log_details)
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

                allmodules= ZohoModules.objects.get(company=company,status='New')
        invoices = invoice.objects.filter(company = company)
        gstIn = request.POST['gstin']

        if Customer.objects.filter(company = company, GST_number__iexact = gstIn).exists():
            msg = f'{gstIn} already exists, Try another.!'
            return JsonResponse({'is_exist':True, 'message':msg})
        else:
            return JsonResponse({'is_exist':False})
    else:
        return redirect('/')
def checkCustomerPAN(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details)
        pan = request.POST['pan']

        if Customer.objects.filter(company = com, PAN_number__iexact = pan).exists():
            msg = f'{pan} already exists, Try another.!'
            return JsonResponse({'is_exist':True, 'message':msg})
        else:
            return JsonResponse({'is_exist':False})
    else:
        return redirect('/')
def checkCustomerPhone(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
               
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details)
        phn = request.POST['phone']

        if Customer.objects.filter(company = com,customer_phone__iexact = phn).exists():
            msg = f'{phn} already exists, Try another.!'
            return JsonResponse({'is_exist':True, 'message':msg})
        else:
            return JsonResponse({'is_exist':False})
    else:
        return redirect('/')
def checkCustomerEmail(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
         
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details)
        email = request.POST['email']

        if Customer.objects.filter(company = com, customer_email__iexact = email).exists():
            msg = f'{email} already exists, Try another.!'
            return JsonResponse({'is_exist':True, 'message':msg})
        else:
            return JsonResponse({'is_exist':False})
    else:
        return redirect('/')
def newCustomerPaymentTerm(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
           
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details)
        term = request.POST['term']
        days = request.POST['days']

        if not Company_Payment_Term.objects.filter(company = com, term_name__iexact = term).exists():
            Company_Payment_Term.objects.create(company = com, term_name = term, days =days)
            
            list= []
            terms = Company_Payment_Term.objects.filter(company = com)

            for term in terms:
                termDict = {
                    'name': term.term_name,
                    'id': term.id,
                    'days':term.days
                }
                list.append(termDict)

            return JsonResponse({'status':True,'terms':list},safe=False)
        else:
            return JsonResponse({'status':False, 'message':f'{term} already exists, try another.!'})

    else:
        return redirect('/')
def add_customer_invoice(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')

        

       
        if request.method=="POST":
            customer_data=Customer()
            customer_data.login_details=log_details
            customer_data.company=comp_details
            customer_data.customer_type = request.POST.get('type')

            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            customer_data.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type=request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type=op_type
            else:
                customer_data.opening_balance_type='Opening Balance not selected'
    
            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            # customer_data.price_list=request.POST['plst']
            plst=request.POST.get('plst')
            if plst!=0:
                 customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'




            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                 customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'
    



           
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=comp_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            # vdata=Customer.objects.get(id=customer_data.id)
            # vendor=vdata
            # rdata=Customer_remarks_table()
            # rdata.remarks=request.POST['remark']
            # rdata.company=comp_details
            # rdata.customer=vdata
            # rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Customer.objects.get(id=customer_data.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = CustomerContactPersons.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype=ele[6],designation=ele[7],department=ele[8],company=comp_details,customer=vendor)
                
        
            messages.success(request, 'Customer created successfully!')   

            return redirect('invoice_createpage')
        
        else:
            messages.error(request, 'Some error occurred !')   

            return redirect('invoice_createpage')
def create_item_invoice(request):                                                                #
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.Date=date.today()
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track = request.POST.get("trackState",None)
            track_state_value = request.POST.get("trackstate", None)

# Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0

            
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if Items.objects.filter(item_name=item_name, company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('invoice_createpage')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track_state_value = request.POST.get("trackState", None)

            track_state_value = request.POST.get("trackstate", None)

            # Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0
            minstock=request.POST.get("minimum_stock",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
        
        

        
            if Items.objects.filter(item_name=item_name,company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('invoice_createpage')
    return redirect('invoice_createpage')

def createInvoiceCustomer(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        if request.method=="POST":
            customer_data=Customer()
            customer_data.login_details=com.login_details
            customer_data.company=com
            customer_data.customer_type = request.POST.get('type')

            customer_data.title = request.POST.get('salutation')
            customer_data.first_name=request.POST['first_name']
            customer_data.last_name=request.POST['last_name']
            customer_data.company_name=request.POST['company_name']
            customer_data.customer_display_name=request.POST['v_display_name']
            customer_data.customer_email=request.POST['vendor_email']
            customer_data.customer_phone=request.POST['w_phone']
            customer_data.customer_mobile=request.POST['m_phone']
            customer_data.skype=request.POST['skype_number']
            customer_data.designation=request.POST['designation']
            customer_data.department=request.POST['department']
            customer_data.website=request.POST['website']
            customer_data.GST_treatement=request.POST['gst']
            customer_data.customer_status="Active"
            customer_data.remarks=request.POST['remark']
            customer_data.current_balance=request.POST['opening_bal']

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                customer_data.PAN_number=request.POST['pan_number']
                customer_data.GST_number="null"
            else:
                customer_data.GST_number=request.POST['gst_number']
                customer_data.PAN_number=request.POST['pan_number']

            customer_data.place_of_supply=request.POST['source_supply']
            customer_data.currency=request.POST['currency']
            op_type = request.POST.get('op_type')
            if op_type is not None:
                customer_data.opening_balance_type = op_type
            else:
                customer_data.opening_balance_type ='Opening Balance not selected'

            customer_data.opening_balance=request.POST['opening_bal']
            customer_data.company_payment_terms= None if request.POST['payment_terms'] == "" else Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            # customer_data.price_list=request.POST['plst']
            plst=request.POST.get('plst')
            if plst!=0:
                    customer_data.price_list=plst
            else:
                customer_data.price_list='Price list not selected'




            # customer_data.portal_language=request.POST['plang']
            plang=request.POST.get('plang')
            if plang!=0:
                    customer_data.portal_language=plang
            else:
                customer_data.portal_language='Portal language not selected'

            customer_data.facebook=request.POST['fbk']
            customer_data.twitter=request.POST['twtr']
            customer_data.tax_preference=request.POST['tax1']

            type=request.POST.get('type')
            if type is not None:
                customer_data.customer_type=type
            else:
                customer_data.customer_type='Customer type not selected'




            
            customer_data.billing_attention=request.POST['battention']
            customer_data.billing_country=request.POST['bcountry']
            customer_data.billing_address=request.POST['baddress']
            customer_data.billing_city=request.POST['bcity']
            customer_data.billing_state=request.POST['bstate']
            customer_data.billing_pincode=request.POST['bzip']
            customer_data.billing_mobile=request.POST['bphone']
            customer_data.billing_fax=request.POST['bfax']
            customer_data.shipping_attention=request.POST['sattention']
            customer_data.shipping_country=request.POST['s_country']
            customer_data.shipping_address=request.POST['saddress']
            customer_data.shipping_city=request.POST['scity']
            customer_data.shipping_state=request.POST['sstate']
            customer_data.shipping_pincode=request.POST['szip']
            customer_data.shipping_mobile=request.POST['sphone']
            customer_data.shipping_fax=request.POST['sfax']
            customer_data.remarks=request.POST['remark']

            customer_data.save()
            
            vendor_history_obj=CustomerHistory()
            vendor_history_obj.company=com
            vendor_history_obj.login_details=com.login_details
            vendor_history_obj.customer=customer_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Completed'
            vendor_history_obj.save()

            # vdata=Customer.objects.get(id=customer_data.id)
            # rdata=Customer_remarks_table()
            # rdata.remarks=request.POST['remark']
            # rdata.company=com
            # rdata.customer=vdata
            # rdata.save()

        
            title =request.POST.getlist('tsalutation[]')
            first_name =request.POST.getlist('tfirstName[]')
            last_name =request.POST.getlist('tlastName[]')
            email =request.POST.getlist('tEmail[]')
            work_phone =request.POST.getlist('tWorkPhone[]')
            mobile =request.POST.getlist('tMobilePhone[]')
            skype_name_number =request.POST.getlist('tSkype[]')
            designation =request.POST.getlist('tDesignation[]')
            department =request.POST.getlist('tDepartment[]') 
            vdata=Customer.objects.get(id=customer_data.id)

            if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                mapped2=list(mapped2)
                print(mapped2)
                for ele in mapped2:
                    CustomerContactPersons.objects.create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],work_phone=ele[4],mobile=ele[5],skype=ele[6],designation=ele[7],department=ele[8],company=com,customer=vdata)
        
            return JsonResponse({'status':True})
        else:
            return JsonResponse({'status':False})
    
def getCustomers(request):
    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        options = {}
        option_objects = Customer.objects.filter(company = com, customer_status = 'Active')
        for option in option_objects:
            options[option.id] = [option.id , option.title, option.first_name, option.last_name]

        return JsonResponse(options)
    else:
        return redirect('/')
def saveItemUnit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
           
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details) 
        if request.method == "POST":
            name = request.POST['units'].upper()
            print(name)

            if not Unit.objects.filter(company = com, unit_name__iexact = name).exists():
                unit = Unit(
                    company = com,
                    unit_name = name
                )
                unit.save()
                return JsonResponse({'status':True})
            else:
                return JsonResponse({'status':False, 'message':'Unit already exists.!'})
def show_unit_dropdown(request):

    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

            options = {}
            option_objects = Unit.objects.filter(company=com)
            for option in option_objects:
                options[option.id] = [option.id,option.unit_name]
            return JsonResponse(options)



def invoice_item(request):   

    if 'login_id' in request.session:
            log_id = request.session['login_id']
            if 'login_id' not in request.session:
                return redirect('/')
            log_details= LoginDetails.objects.get(id=log_id)
                
            if log_details.user_type == 'Staff':
                    staff = StaffDetails.objects.get(login_details=log_details)
                    com = staff.company
                        
            elif log_details.user_type == 'Company':
                    com = CompanyDetails.objects.get(login_details=log_details) 
            if request.method=='POST':
                
                type=request.POST.get('type')
                
                name=request.POST.get('name')

                ut=request.POST.get('unit')
                inter=request.POST.get('inter')
                intra=request.POST.get('intra')
                sell_price=request.POST.get('sell_price')
                sell_acc=request.POST.get('sell_acc')
                sell_desc=request.POST.get('sell_desc')
                cost_price=request.POST.get('cost_price')
                cost_acc=request.POST.get('cost_acc')      
                cost_desc=request.POST.get('cost_desc')
                hsn_number = request.POST.get('hsn_number')
                stock = 0 if request.POST.get('stock') == "" else request.POST.get('stock')
                stockUnitRate = 0 if request.POST.get('stock_rate') == "" else request.POST.get('stock_rate')
                print(stock)

                print(stockUnitRate)

                

                
                
                units=Unit.objects.get(id=ut)
                

                if Items.objects.filter(company=com, item_name__iexact=name).exists():
                    res = f"{name} already exists, try another!"
                    return JsonResponse({'status': False, 'message':res})
                elif Items.objects.filter(company = com, hsn_code__iexact = hsn_number).exists():
                    res = f"HSN - {hsn_number} already exists, try another.!"
                    return JsonResponse({'status': False, 'message':res})
                else:


                    item=Items(company = com,
                        login_details = log_details,
                        item_name = name,
                        item_type = type,
                        unit = units,
                        hsn_code = hsn_number,
                        intrastate_tax = intra,
                        interstate_tax = inter,
                        sales_account = sell_acc,
                        selling_price = sell_price,
                        sales_description = sell_desc,
                        purchase_account = cost_acc,
                        purchase_price = cost_price,
                        purchase_description = cost_desc,
                        # date = createdDate,
                        # inventory_account = inventory,
                        opening_stock = stock,
                        current_stock = stock,
                       
                        opening_stock_per_unit = stockUnitRate,
                        activation_tag = 'active'
        )

                item.save()
                Item_Transaction_History.objects.create(
                company = com,
                logindetails = log_details,
                items = item,
                action = 'Created'
            )
            

                return HttpResponse({"message": "success"})
    
    return HttpResponse("Invalid request method.")
def createInvoiceItem(request):

   
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        name = request.POST['name']
        type = request.POST['type']
        unit = request.POST.get('unit')
        hsn = request.POST['hsn']
        tax = request.POST['taxref']
        gstTax = 0 if tax == 'None-Taxable' else request.POST['intra_st']
        igstTax = 0 if tax == 'None-Taxable' else request.POST['inter_st']
        purPrice = request.POST['pcost']
        purAccount = request.POST['pur_account']
        purDesc = request.POST['pur_desc']
        salePrice = request.POST['salesprice']
        saleAccount = request.POST['sale_account']
        saleDesc = request.POST['sale_desc']
        inventory = request.POST.get('invacc')
        stock = 0 if request.POST.get('stock') == "" else request.POST.get('stock')
        stockUnitRate = 0 if request.POST.get('stock_rate') == "" else request.POST.get('stock_rate')
        minStock = request.POST['min_stock']
        createdDate = date.today()
        print("ok")
        
        #save item and transaction if item or hsn doesn't exists already
        if Items.objects.filter(company=com, item_name__iexact=name).exists():
            res = f"{name} already exists, try another!"
            return JsonResponse({'status': False, 'message':res})
        elif Items.objects.filter(company = com, hsn_code__iexact = hsn).exists():
            res = f"HSN - {hsn} already exists, try another.!"
            return JsonResponse({'status': False, 'message':res})
        else:
            item = Items(
                company = com,
                login_details = com.login_details,
                item_name = name,
                item_type = type,
                unit = None if unit == "" else Unit.objects.get(id = int(unit)),
                hsn_code = hsn,
                tax_reference = tax,
                intrastate_tax = gstTax,
                interstate_tax = igstTax,
                sales_account = saleAccount,
                selling_price = salePrice,
                sales_description = saleDesc,
                purchase_account = purAccount,
                purchase_price = purPrice,
                purchase_description = purDesc,
                date = createdDate,
                minimum_stock_to_maintain = minStock,
                inventory_account = inventory,
                opening_stock = stock,
                current_stock = stock,
                opening_stock_per_unit = stockUnitRate,
                track_inventory = int(request.POST['trackInv']),
                activation_tag = 'active',
                type = 'Opening Stock'
            )
            item.save()

            #save transaction

            Item_Transaction_History.objects.create(
                company = com,
                logindetails = com.login_details,
                items = item,
                Date = createdDate,
                action = 'Created'

            )
            
            return JsonResponse({'status': True})
    else:
       return redirect('/')
def getItems(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        items = {}
        option_objects = Items.objects.filter(company = com, activation_tag='active')
        for option in option_objects:
            items[option.id] = [option.id,option.item_name]

        return JsonResponse(items)
    else:
        return redirect('/')
    
def checkAccounts(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
            
        if log_details.user_type == 'Staff':
                staff = StaffDetails.objects.get(login_details=log_details)
                com = staff.company
                    
        elif log_details.user_type == 'Company':
                com = CompanyDetails.objects.get(login_details=log_details)

        if Chart_of_Accounts.objects.filter(company = com, account_type = request.GET['type']).exists():
            list= []
            account_objects = Chart_of_Accounts.objects.filter(company = com, account_type = request.GET['type'])

            for account in account_objects:
                accounts = {
                    'name': account.account_name,
                }
                list.append(accounts)

            return JsonResponse({'status':True,'accounts':list},safe=False)
        else:
            return JsonResponse({'status':False})
def createNewAccountFromItems(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            b.company=com
            b.logindetails=com.login_details
            b.action="Created"
            b.Date=date.today()
            a.login_details=com.login_details
            a.company=com
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number",None)
            a.account_description = request.POST['description']
            a.sub_account = request.POST.get("sub_acc",None)
            a.parent_account = request.POST.get("parent_acc",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
    
            a.Create_status="added"
            a.status = 'Active'
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=com).exists():
                return JsonResponse({'status': False, 'message':'Account Name already exists.!'})
            else:
                a.save()
                b.chart_of_accounts=a
                b.save()
                return JsonResponse({'status': True})

    else:
        return redirect('/')

def getAllAccounts(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        acc = {}
        acc_objects = Chart_of_Accounts.objects.filter(company = com, status = 'Active')
        for option in acc_objects:
            acc[option.id] = [option.account_name,option.account_type]

        return JsonResponse(acc)
    else:
        return redirect('/')
       
def company_gsttype_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            
            gstno = request.POST.get('gstno')
            gsttype = request.POST.get('gsttype')

            # Check if gsttype is one of the specified values
            if gsttype in ['unregistered Business', 'Overseas', 'Consumer']:
                dash_details.gst_no = None
            else:
                if gstno:
                    dash_details.gst_no = gstno
                else:
                    messages.error(request,'GST Number is not entered*')
                    return redirect('company_profile_editpage')


            dash_details.gst_type = gsttype

            dash_details.save()
            messages.success(request,'GST Type changed')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/') 
    

# -------------------------------Zoho Modules section--------------------------------

def check_customer_phonenumber_exist(request):
    if request.method == 'GET':
       mPhone = request.GET.get('m_Phone', None)

       if mPhone:
          
            exists = Customer.objects.filter(
                    customer_mobile=mPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def check_customer_work_phone_exist(request):
    if request.method == 'GET':
       wPhone = request.GET.get('w_Phone', None)

       if wPhone:
          
            exists = Customer.objects.filter(
                    customer_phone=wPhone
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})   

def check_customer_email_exist(request):
    if request.method == 'GET':
       vendoremail = request.GET.get('vendor_email', None)

       if vendoremail:
          
            exists = Customer.objects.filter(
                    customer_email=vendoremail
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False}) 

def customer_payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
    
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            normalized_data = terms.replace(" ", "")
            pay_tm = add_space_before_first_digit(normalized_data)
            ptr = Company_Payment_Term(term_name=pay_tm, days=day, company=dash_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=comp_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)   
            

def check_customer_term_exist(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(id=dash_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            comp_details=CompanyDetails.objects.get(login_details=log_details)

            
        allmodules= ZohoModules.objects.get(company=comp_details,status='New')
   

    if request.method == 'GET':
       term_name = request.GET.get('term_name', None)
       if term_name:
            normalized_data = term_name.replace(" ", "")
            term_name_processed = add_space_before_first_digit(normalized_data)
            exists = Company_Payment_Term.objects.filter(
                    term_name=term_name_processed,
                    company=comp_details
                ).exists()
            return JsonResponse({'exists': exists})          
    else:
        return JsonResponse({'exists': False})    

def customer_check_pan(request):
    if request.method == 'POST':
        panNumber = request.POST.get('panNumber')
        pan_exists = Customer.objects.filter(PAN_number=panNumber).exists()

        if pan_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'})  

def customer_check_gst(request):
    if request.method == 'POST':
        gstNumber = request.POST.get('gstNumber')
        gst_exists = Customer.objects.filter(GST_number=gstNumber).exists()
       
        if gst_exists:
            return JsonResponse({'status': 'exists'})
        else:
            return JsonResponse({'status': 'not_exists'})
    else:
        return JsonResponse({'error': 'Invalid request'})
    
    # edited
def getinvCustomerDetails(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company
        
        custId = request.POST['id']
        cust = Customer.objects.get(id = custId)

        if cust:
            context = {
                'status':True, 'id':cust.id, 'email':cust.customer_email, 'gstType':cust.GST_treatement,'shipState':cust.place_of_supply,'gstin':False if cust.GST_number == "" or cust.GST_number == None else True, 'gstNo':cust.GST_number,
                'street':cust.billing_address, 'city':cust.billing_city, 'state':cust.billing_state, 'country':cust.billing_country, 'pincode':cust.billing_pincode
            }
            return JsonResponse(context)
        else:
            return JsonResponse({'status':False, 'message':'Something went wrong..!'})
    else:
       return redirect('/')
def getinvBankAccountNumber(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company
        
        bankId = request.GET['id']
        bnk = Banking.objects.get(id = bankId)

        if bnk:
            return JsonResponse({'status':True, 'account':bnk.bnk_acno})
        else:
            return JsonResponse({'status':False, 'message':'Something went wrong..!'})
    else:
       return redirect('/')
def newinvPaymentTerm(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        term = request.POST['term']
        days = request.POST['days']

        if not Company_Payment_Term.objects.filter(company = com, term_name__iexact = term).exists():
            Company_Payment_Term.objects.create(company = com, term_name = term, days =days)
            
            list= []
            terms = Company_Payment_Term.objects.filter(company = com)

            for term in terms:
                termDict = {
                    'name': term.term_name,
                    'id': term.id,
                    'days':term.days
                }
                list.append(termDict)

            return JsonResponse({'status':True,'terms':list},safe=False)
        else:
            return JsonResponse({'status':False, 'message':f'{term} already exists, try another.!'})

    else:
        return redirect('/')
def getinvItemDetails(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company
        
        itemName = request.GET['item']
        priceListId = request.GET['listId']
        item = Items.objects.filter(company = cmp, item_name = itemName).first()

        if priceListId != "":
            priceList = PriceList.objects.get(id = int(priceListId))

            if priceList.item_rate_type == 'Each Item':
                try:
                    priceListPrice = float(PriceListItem.objects.get(company = cmp, price_list = priceList, item = item).custom_rate)
                except:
                    priceListPrice = item.selling_price
            else:
                mark = priceList.percentage_type
                percentage = float(priceList.percentage_value)
                roundOff = priceList.round_off

                if mark == 'Markup':
                    price = float(item.selling_price) + float((item.selling_price) * (percentage/100))
                else:
                    price = float(item.selling_price) - float((item.selling_price) * (percentage/100))

                if priceList.round_off != 'Never Mind':
                    if roundOff == 'Nearest Whole Number':
                        finalPrice = round(price)
                    else:
                        finalPrice = int(price) + float(roundOff)
                else:
                    finalPrice = price

                priceListPrice = finalPrice
        else:
            priceListPrice = None

        context = {
            'status':True,
            'id': item.id,
            'hsn':item.hsn_code,
            'sales_rate':item.selling_price,
            'purchase_rate':item.purchase_price,
            'avl':item.current_stock,
            'tax': True if item.tax_reference == 'taxable' else False,
            'gst':item.intrastate_tax,
            'igst':item.interstate_tax,
            'PLPrice':priceListPrice,

        }
        return JsonResponse(context)
    else:
       return redirect('/')
def checkInvoiceNumber(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company
        
        RecInvNo = request.GET['RecInvNum']

        # Finding next rec_invoice number w r t last rec_invoice number if exists.
        nxtInv = ""
        lastInv = invoice.objects.filter(company = com).last()

        if lastInv:
            inv_no = str(lastInv.invoice_number)

            numbers = []
            stri = []
            for word in inv_no:
                if word.isdigit():
                    numbers.append(word)
                else:
                    stri.append(word)
            
            num=''
            for i in numbers:
                num +=i
            
            st = ''
            for j in stri:
                st = st+j            


            inv_num = int(num)+1
           

            
            padding_length = len(num) - 1

                    
            nxtInv = f"{st}{num[0]}{inv_num:0{padding_length}d}"
            print(nxtInv)
            

        PatternStr = []
        for word in RecInvNo:
            if word.isdigit():
                pass
            else:
                PatternStr.append(word)
        
        pattern = ''
        for j in PatternStr:
            pattern += j
            print("patern")
            print(pattern)


        # pattern_exists = checkRecInvNumberPattern(pattern)

        # if pattern !="" and pattern_exists:
        #     return JsonResponse({'status':False, 'message':'Rec. Invoice No. Pattern already Exists.!'})
        if invoice.objects.filter(company = com, invoice_number__iexact = RecInvNo).exists():
            return JsonResponse({'status':False, 'message':'Invoice No. already Exists.!'})
        elif nxtInv != "" and RecInvNo != nxtInv:
            return JsonResponse({'status':False, 'message':'Invoice No. is not continuous.!'})
        else:
            return JsonResponse({'status':True, 'message':'Number is okay.!'})
    else:
       return redirect('/')
def addinv_unit(request):                                                               
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        if request.method == 'POST':
            c = CompanyDetails.objects.get(login_details=login_id)
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    elif log_user.user_type == 'Staff':
        if request.method == 'POST':
            staff = LoginDetails.objects.get(id=login_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c = sf.company
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    return JsonResponse({"message": "success"})
# create unit


    
def showinvunit_dropdown(request):                                                               
   if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

            options = {}
            option_objects = Unit.objects.filter(company=com)
            for option in option_objects:
                options[option.id] = [option.id,option.unit_name]
            return JsonResponse(options)
def createNewIteminv(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        name = request.POST['name']
        type = request.POST['type']
        unit = request.POST.get('unit')
        hsn = request.POST['hsn']
        tax = request.POST['taxref']
        gstTax = 0 if tax == 'None-Taxable' else request.POST['intra_st']
        igstTax = 0 if tax == 'None-Taxable' else request.POST['inter_st']
        purPrice = request.POST['pcost']
        purAccount = None if not 'pur_account' in request.POST or request.POST['pur_account'] == "" else request.POST['pur_account']
        purDesc = request.POST['pur_desc']
        salePrice = request.POST['salesprice']
        saleAccount = None if not 'sale_account' in request.POST or request.POST['sale_account'] == "" else request.POST['sale_account']
        saleDesc = request.POST['sale_desc']
        inventory = request.POST.get('invacc')
        stock = 0 if request.POST.get('stock') == "" else request.POST.get('stock')
        stockUnitRate = 0 if request.POST.get('stock_rate') == "" else request.POST.get('stock_rate')
        minStock = request.POST['min_stock']
        createdDate = date.today()
        
        #save item and transaction if item or hsn doesn't exists already
        if Items.objects.filter(company=com, item_name__iexact=name).exists():
            res = f"{name} already exists, try another!"
            return JsonResponse({'status': False, 'message':res})
        elif Items.objects.filter(company = com, hsn_code__iexact = hsn).exists():
            res = f"HSN - {hsn} already exists, try another.!"
            return JsonResponse({'status': False, 'message':res})
        else:
            item = Items(
                company = com,
                login_details = com.login_details,
                item_name = name,
                item_type = type,
                unit = None if unit == "" else Unit.objects.get(id = int(unit)),
                hsn_code = hsn,
                tax_reference = tax,
                intrastate_tax = gstTax,
                interstate_tax = igstTax,
                sales_account = saleAccount,
                selling_price = salePrice,
                sales_description = saleDesc,
                purchase_account = purAccount,
                purchase_price = purPrice,
                purchase_description = purDesc,
                date = createdDate,
                minimum_stock_to_maintain = minStock,
                inventory_account = inventory,
                opening_stock = stock,
                current_stock = stock,
                opening_stock_per_unit = stockUnitRate,
                track_inventory = int(request.POST['trackInv']),
                activation_tag = 'active',
                type = 'Opening Stock'
            )
            item.save()

            #save transaction

            Item_Transaction_History.objects.create(
                company = com,
                logindetails = com.login_details,
                items = item,
                Date = createdDate,
                action = 'Created'

            )
            
            return JsonResponse({'status': True})
    else:
       return redirect('/')

def getAllItemsinv(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            com = CompanyDetails.objects.get(login_details = log_details)
        else:
            com = StaffDetails.objects.get(login_details = log_details).company

        items = {}
        option_objects = Items.objects.filter(company = com, activation_tag='active')
        for option in option_objects:
            items[option.id] = [option.id,option.item_name]
        print(items)

        return JsonResponse(items)
    
    else:
        return redirect('/')

